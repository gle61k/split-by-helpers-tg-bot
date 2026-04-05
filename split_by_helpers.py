from pathlib import Path
from typing import Optional

import pandas as pd

# ================== НАСТРОЙКИ ==================
SCRIPT_DIR = Path(__file__).resolve().parent

COL_HELPER = "Помощник ТУ"
COL_GROUP = "Группа помощника"
COL_HOURS = "Всего Часов"

GROUP_RULES = {
    "Группа 1": "УР_Группа_1.xlsx",
    "Группа 2": "УР_Группа_2.xlsx",
    "Группа 3": "УР_Группа_3.xlsx",
    "Группа 4": "УР_Группа_4.xlsx",
}

OTHER_FILE = "УР_Остальные.xlsx"
# ===============================================


def find_column(df, *candidates):
    """Ищет колонку по нескольким возможным названиям."""
    cols = list(df.columns)
    for c in candidates:
        if c in cols:
            return c
    for c in candidates:
        for col in cols:
            if c in str(col) or c.replace(" ", "") in str(col).replace(" ", ""):
                return col
    for col in cols:
        for c in candidates:
            if c.lower() in str(col).lower():
                return col
    return None


def format_sheet(ws):
    from openpyxl.styles import Font, PatternFill

    header_fill = PatternFill("solid", fgColor="BDD7EE")
    header_font = Font(bold=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)


def save_group(df_group, filename, col_helper, col_hours, add_no_helper=False, df_no_helper=None):
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        summary = []

        helpers = sorted(
            df_group[col_helper]
            .dropna()
            .astype(str)
            .str.strip()
            .unique()
        )

        for helper in helpers:
            h_df = df_group[df_group[col_helper] == helper]
            summary.append({
                "Помощник": helper,
                "Количество сотрудников": len(h_df),
                "Сумма часов": h_df[col_hours].sum()
            })

        summary_df = pd.DataFrame(summary).sort_values("Помощник")
        summary_df.to_excel(writer, sheet_name="Сводный", index=False)

        for helper in helpers:
            h_df = df_group[df_group[col_helper] == helper]
            h_df.to_excel(writer, sheet_name=helper[:31], index=False)

        if add_no_helper and df_no_helper is not None and not df_no_helper.empty:
            df_no_helper.to_excel(writer, sheet_name="Без помощника", index=False)

        wb = writer.book
        for ws in wb.worksheets:
            format_sheet(ws)


def find_export_file():
    """Ищет файл выгрузки в папке скрипта."""
    candidates = list(SCRIPT_DIR.glob("*выгрузка*.xls")) + list(SCRIPT_DIR.glob("*выгрузка*.xlsx"))
    if not candidates:
        raise FileNotFoundError(f"В папке {SCRIPT_DIR} не найден файл выгрузки (*выгрузка*.xls)")
    return str(max(candidates, key=lambda x: x.stat().st_mtime))


def _read_excel(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix == ".xls":
        return pd.read_excel(path, engine="xlrd")
    return pd.read_excel(path, engine="openpyxl")


def process_export(
    export_path: str | Path,
    output_dir: Optional[Path] = None,
) -> tuple[list[Path], list[str]]:
    """
    Читает выгрузку, пишет xlsx в output_dir (по умолчанию — папка скрипта).
    Возвращает (список созданных файлов, текстовые заметки для пользователя).
    """
    export_path = Path(export_path)
    out = Path(output_dir) if output_dir is not None else SCRIPT_DIR
    out.mkdir(parents=True, exist_ok=True)
    notes: list[str] = []
    created: list[Path] = []

    df = _read_excel(export_path)

    col_helper = find_column(df, "Помощник ТУ", "ПомощникТУ", "Помощник")
    col_group = find_column(df, "Группа помощника", "ГруппаПомощника", "Группа")
    col_hours = find_column(df, "Всего Часов", "ВсегоЧасов", "Часы")
    if not col_hours:
        for c in df.columns:
            if "час" in str(c).lower():
                col_hours = c
                break
    if not col_helper:
        raise ValueError(
            "Не найдена колонка помощника (Помощник ТУ / Помощник). Колонки в файле: "
            + ", ".join(repr(x) for x in df.columns)
        )
    if not col_hours:
        raise ValueError(
            "Не найдена колонка часов. Колонки в файле: " + ", ".join(repr(x) for x in df.columns)
        )

    df = df.copy()
    df[col_helper] = df[col_helper].astype(str).str.strip()
    df[col_hours] = pd.to_numeric(df[col_hours], errors="coerce").fillna(0)

    bad_helper_mask = df[col_helper].isin(["", "nan", "None"]) | df[col_helper].isna()
    df_no_helper = df[bad_helper_mask]

    if not col_group:
        notes.append(
            "Колонка группы не найдена — весь объём попадает в «УР_Остальные.xlsx» "
            "(файлы по группам 1–4 не формируются)."
        )
        other_df = df
        dest = out / OTHER_FILE
        save_group(
            other_df[~bad_helper_mask],
            str(dest),
            col_helper,
            col_hours,
            add_no_helper=True,
            df_no_helper=df_no_helper,
        )
        created.append(dest)
        return created, notes

    df[col_group] = df[col_group].astype(str).str.strip()
    used_idx: set = set()

    for group_key, out_file in GROUP_RULES.items():
        group_df = df[df[col_group].str.contains(group_key, na=False)]

        if group_df.empty:
            notes.append(f"{out_file} — нет данных")
            continue

        used_idx.update(group_df.index)
        dest = out / out_file
        save_group(group_df, str(dest), col_helper, col_hours)
        created.append(dest)

    other_df = df.loc[~df.index.isin(used_idx)]
    dest_other = out / OTHER_FILE
    save_group(
        other_df[other_df[col_helper].notna()],
        str(dest_other),
        col_helper,
        col_hours,
        add_no_helper=True,
        df_no_helper=df_no_helper,
    )
    created.append(dest_other)
    return created, notes


def main():
    export_path = find_export_file()
    print(f"Читаю файл: {Path(export_path).name}")
    try:
        created, notes = process_export(export_path)
    except ValueError as e:
        raise SystemExit(str(e)) from e
    for line in notes:
        print(line)
    for p in created:
        print(f"создан {p.name}")
    print("ГОТОВО ✅")


if __name__ == "__main__":
    main()
