# -*- coding: utf-8 -*-
"""
Разбивка выгрузки часов по группам помощников.
Группа определяется по справочнику PR.xlsx (ФИО → группа).
Запуск: положите скрипт в папку с PR.xlsx и файлом выгрузки, запустите.
"""
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

# --- Пути (скрипт сам ищет файлы в своей папке) ---
SCRIPT_DIR = Path(__file__).resolve().parent
PR_FILENAME = "PR.xlsx"
EXPORT_PATTERN = "*выгрузка*.xls"  # или *выгрузка*.xlsx
OUTPUT_FILES = {
    "1": "УР_Группа_1.xlsx",
    "2": "УР_Группа_2.xlsx",
    "3": "УР_Группа_3.xlsx",
    "4": "УР_Группа_4.xlsx",
    "Остальные": "УР_Остальные.xlsx",
}


def find_pr_path():
    p = SCRIPT_DIR / PR_FILENAME
    if p.exists():
        return str(p)
    for f in SCRIPT_DIR.glob("*PR*.xlsx"):
        return str(f)
    raise FileNotFoundError(f"В папке {SCRIPT_DIR} не найден файл {PR_FILENAME}")


def find_export_path():
    candidates = list(SCRIPT_DIR.glob("*выгрузка*.xls")) + list(SCRIPT_DIR.glob("*выгрузка*.xlsx"))
    if not candidates:
        raise FileNotFoundError(f"В папке {SCRIPT_DIR} не найден файл выгрузки (*выгрузка*.xls)")
    return str(max(candidates, key=lambda x: x.stat().st_mtime))


def parse_group_from_text(group_text):
    """Из текста группы (например 'ПУ. Группа 1') возвращает '1', '2', '3', '4' или 'Остальные'."""
    if pd.isna(group_text) or not str(group_text).strip():
        return "Остальные"
    s = str(group_text).strip()
    if re.search(r"группа\s*1\b", s, re.IGNORECASE) or (re.search(r"\b1\b", s) and "группа" in s.lower()):
        return "1"
    if re.search(r"группа\s*2\b", s, re.IGNORECASE) or (re.search(r"\b2\b", s) and "группа" in s.lower()):
        return "2"
    if re.search(r"группа\s*3\b", s, re.IGNORECASE) or (re.search(r"\b3\b", s) and "группа" in s.lower()):
        return "3"
    if re.search(r"группа\s*4\b", s, re.IGNORECASE) or (re.search(r"\b4\b", s) and "группа" in s.lower()):
        return "4"
    return "Остальные"


def _normalize_name(s):
    """Убирает лишние пробелы для надёжного сопоставления ФИО."""
    if pd.isna(s) or not str(s).strip():
        return ""
    return " ".join(str(s).strip().split())


def load_pr(pr_path):
    """
    Читает PR.xlsx. Структура: каждая КОЛОНКА = одна группа.
    Заголовок колонки = название группы (ПУ. Группа 1, ПУ. Группа 2, ...),
    под ним — список ФИО помощников этой группы.
    Возвращает словарь: нормализованное ФИО -> '1'|'2'|'3'|'4'|'Остальные'.
    """
    df = pd.read_excel(pr_path, sheet_name=0, header=0)
    result = {}
    for col in df.columns:
        header = str(col).strip()
        group_num = parse_group_from_text(header)
        for val in df[col].dropna().astype(str).str.strip():
            if val and val.lower() not in ("nan", "none", ""):
                result[_normalize_name(val)] = group_num
    return result


def find_column(df, *candidates):
    """Возвращает имя колонки из df, если она есть или содержится в каком-то названии."""
    cols = list(df.columns)
    for c in candidates:
        if c in cols:
            return c
    for c in candidates:
        for col in cols:
            if c in str(col) or c.replace(" ", "") in str(col).replace(" ", ""):
                return col
    return None


def load_export(export_path):
    """Читает выгрузку, находит колонки, добавляет 'Группа помощника'. Возвращает DataFrame."""
    df = pd.read_excel(export_path)
    # Нужные колонки
    col_helper = find_column(df, "Помощник ТУ", "ПомощникТУ", "Помощник")
    col_hours = find_column(df, "ВсегоЧасов", "Всего Часов", "Часы")
    if not col_hours:
        for c in df.columns:
            if "час" in str(c).lower():
                col_hours = c
                break
    if not col_helper:
        raise SystemExit("В выгрузке не найдена колонка помощника (Помощник ТУ / ПомощникТУ).")
    if not col_hours:
        raise SystemExit("В выгрузке не найдена колонка часов (Часы / ВсегоЧасов).")

    df[col_helper] = df[col_helper].astype(str).str.strip()
    df[col_hours] = pd.to_numeric(df[col_hours], errors="coerce").fillna(0)
    return df, col_helper, col_hours


def add_group_column(df, col_helper, pr_dict):
    """Добавляет колонку 'Группа помощника' по PR."""
    def group_for_row(helper):
        if not helper or str(helper).strip() in ("nan", "None", ""):
            return "Остальные"
        return pr_dict.get(_normalize_name(helper), "Остальные")

    df["Группа помощника"] = df[col_helper].map(group_for_row)
    return df


def split_by_groups(df, col_helper, col_hours):
    """Разбивает DataFrame по группам. Возвращает dict: группа -> (df с помощниками, df без помощника)."""
    no_helper = df[df[col_helper].isin(("", "nan", "None")) | df[col_helper].isna()]
    with_helper = df[~df.index.isin(no_helper.index)]

    result = {}
    for group_key in ("1", "2", "3", "4", "Остальные"):
        group_df = with_helper[with_helper["Группа помощника"] == group_key].copy()
        result[group_key] = (group_df, no_helper if group_key == "Остальные" else pd.DataFrame())
    return result


def build_summary_df(group_df, col_helper, col_hours):
    """Сводный лист: № п/п, Помощник, Количество сотрудников, Сумма часов. Сортировка по алфавиту."""
    helpers = (
        group_df[col_helper]
        .dropna()
        .astype(str)
        .str.strip()
        .replace({"": pd.NA, "nan": pd.NA, "None": pd.NA})
        .dropna()
        .unique()
    )
    helpers = sorted(helpers)
    rows = []
    for i, h in enumerate(helpers, 1):
        sub = group_df[group_df[col_helper].astype(str).str.strip() == h]
        rows.append({
            "№ п/п": i,
            "Помощник": h,
            "Количество сотрудников": len(sub),
            "Сумма часов": sub[col_hours].sum(),
        })
    return pd.DataFrame(rows)


def apply_format(ws, header_fill, header_font, thin_border):
    """Заголовки: жирный, светло-синий фон. Автоширина. Границы у ячеек."""
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    for col in ws.columns:
        if col[0].column_letter:
            max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)


def save_group_file(group_key, group_df, no_helper_df, col_helper, col_hours, out_path):
    """Один выходной файл: Сводный, листы по помощникам (по алфавиту), в конце 'Без помощника' при необходимости."""
    header_fill = PatternFill("solid", fgColor="BDD7EE")
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # 1) Сводный — всегда первый
        summary = build_summary_df(group_df, col_helper, col_hours)
        summary.to_excel(writer, sheet_name="Сводный", index=False)

        # 2) Листы по помощникам, по алфавиту
        helpers = (
            group_df[col_helper]
            .dropna()
            .astype(str)
            .str.strip()
            .replace({"": pd.NA, "nan": pd.NA, "None": pd.NA})
            .dropna()
            .unique()
        )
        helpers = sorted(helpers)
        for h in helpers:
            sub = group_df[group_df[col_helper].astype(str).str.strip() == h]
            sheet_name = (str(h)[:31]).replace("*", "").replace(":", "").replace("?", "").replace("[", "").replace("]", "")
            if not sheet_name:
                sheet_name = "Помощник"
            sub.to_excel(writer, sheet_name=sheet_name, index=False)

        # 3) Без помощника — последний лист (только для Остальные и если есть данные)
        if not no_helper_df.empty:
            no_helper_df.to_excel(writer, sheet_name="Без помощника", index=False)

    wb = load_workbook(out_path)
    for ws in wb.worksheets:
        apply_format(ws, header_fill, header_font, thin_border)
    wb.save(out_path)


def main():
    print("Поиск файлов...")
    pr_path = find_pr_path()
    export_path = find_export_path()
    print(f"  PR: {Path(pr_path).name}")
    print(f"  Выгрузка: {Path(export_path).name}")

    print("Читаю PR...")
    pr_dict = load_pr(pr_path)
    print(f"  Загружено {len(pr_dict)} записей (ФИО → группа).")

    print("Читаю выгрузку...")
    df, col_helper, col_hours = load_export(export_path)
    print(f"  Колонка помощника: {col_helper}, колонка часов: {col_hours}")

    df = add_group_column(df, col_helper, pr_dict)

    splits = split_by_groups(df, col_helper, col_hours)

    for group_key, out_name in OUTPUT_FILES.items():
        group_df, no_helper_df = splits[group_key]
        out_path = SCRIPT_DIR / out_name
        if group_df.empty and no_helper_df.empty and group_key != "Остальные":
            print(f"  {out_name} — нет данных, пропуск.")
            continue
        save_group_file(group_key, group_df, no_helper_df, col_helper, col_hours, str(out_path))
        print(f"  Создан {out_name}")

    print("ГОТОВО.")


if __name__ == "__main__":
    main()
